from itertools import groupby
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from app.models.aplicacao import DemonstrativoCarteira, LinhaCarteira
from app.utils.formatadores import cnpj, data_curta


CABECALHOS = [
    "PRODUTO", "BANCO", "TIPO", "DATA EMISSAO", "DATA VCTO",
    "PRAZO", "TAXA", "VALOR DA APLICACAO", "RENDIMENTO BRUTO (%)",
    "VALOR ATUALIZADO", "RENDIMENTO BRUTO (R$)", "IR", "IOF", "RESGATE LIQUIDO",
]

LARGURAS = [40, 20, 18, 14, 14, 10, 18, 20, 20, 20, 20, 14, 14, 20]
FORMATO_MOEDA = '#,##0.00'
FORMATO_PERCENTUAL = '0.00%'
COLUNAS_MOEDA = [8, 10, 11, 12, 13, 14]
COLUNA_PERCENTUAL = 9

COR_CABECALHO_EMPRESA = "4472C4"
COR_CABECALHO_TABELA = "BFBFBF"
COR_TOTAIS = "E6E6E6"


class RelatorioExcelCarteira:
    def __init__(self) -> None:
        self.pasta_saida = Path("data/excel")
        self.pasta_saida.mkdir(parents=True, exist_ok=True)

    def gerar(self, demonstrativo: DemonstrativoCarteira) -> Path:
        caminho = self.pasta_saida / f"carteira_{demonstrativo.data_saldo.isoformat()}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Saldo da Carteira"

        grupos = self._agrupar_por_empresa(demonstrativo.carteira)
        for empresa_nome, empresa_cnpj, itens in grupos:
            if empresa_nome:
                rotulo = empresa_nome
                if empresa_cnpj:
                    rotulo += f" — CNPJ: {cnpj(empresa_cnpj)}"
                self._escrever_cabecalho_empresa(ws, rotulo)

            self._escrever_cabecalho_tabela(ws)
            self._escrever_linhas(ws, itens)
            self._escrever_totais(ws, itens)
            ws.append([])

        for i, largura in enumerate(LARGURAS, 1):
            ws.column_dimensions[get_column_letter(i)].width = largura

        wb.save(caminho)
        return caminho

    def _agrupar_por_empresa(self, carteira: list[LinhaCarteira]) -> list[tuple]:
        ordenado = sorted(carteira, key=lambda x: (x.empresa_nome or "", x.produto))
        resultado = []
        for (nome, cnpj_val), grupo in groupby(ordenado, key=lambda x: (x.empresa_nome, x.empresa_cnpj)):
            resultado.append((nome, cnpj_val, list(grupo)))
        return resultado

    def _escrever_cabecalho_empresa(self, ws, rotulo: str) -> None:
        ws.append([rotulo])
        linha = ws.max_row
        cell = ws.cell(linha, 1)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COR_CABECALHO_EMPRESA)
        cell.alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(CABECALHOS))

    def _escrever_cabecalho_tabela(self, ws) -> None:
        ws.append(CABECALHOS)
        linha = ws.max_row
        fill = PatternFill("solid", fgColor=COR_CABECALHO_TABELA)
        for cell in ws[linha]:
            cell.font = Font(bold=True)
            cell.fill = fill

    def _escrever_linhas(self, ws, itens: list[LinhaCarteira]) -> None:
        for item in itens:
            ws.append([
                item.produto,
                item.banco,
                item.tipo,
                data_curta(item.data_emissao),
                data_curta(item.data_vencimento),
                item.prazo,
                item.taxa,
                float(item.valor_aplicacao),
                float(item.rendimento_bruto_percentual) / 100,
                float(item.valor_atualizado),
                float(item.rendimento_bruto),
                float(item.valor_ir) if item.valor_ir > 0 else None,
                float(item.valor_iof) if item.valor_iof > 0 else None,
                float(item.resgate_liquido),
            ])
            linha = ws.max_row
            for col in COLUNAS_MOEDA:
                ws.cell(linha, col).number_format = FORMATO_MOEDA
            ws.cell(linha, COLUNA_PERCENTUAL).number_format = FORMATO_PERCENTUAL

    def _escrever_totais(self, ws, itens: list[LinhaCarteira]) -> None:
        ws.append([
            "TOTAIS", "", "", "", "", "", "",
            float(sum(i.valor_aplicacao for i in itens)), "",
            float(sum(i.valor_atualizado for i in itens)),
            float(sum(i.rendimento_bruto for i in itens)),
            float(sum(i.valor_ir for i in itens)),
            float(sum(i.valor_iof for i in itens)),
            float(sum(i.resgate_liquido for i in itens))
        ])
        linha = ws.max_row
        fill = PatternFill("solid", fgColor=COR_TOTAIS)
        for cell in ws[linha]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for col in COLUNAS_MOEDA:
            ws.cell(linha, col).number_format = FORMATO_MOEDA
